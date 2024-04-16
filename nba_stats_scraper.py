#    _   _ ____    _      ____                       _             
#   | \ | | __ )  / \    / ___|  ___ _ __ __ _ _ __ (_)_ __   __ _ 
#   |  \| |  _ \ / _ \   \___ \ / __| '__/ _` | '_ \| | '_ \ / _` |
#   | |\  | |_) / ___ \   ___) | (__| | | (_| | |_) | | | | | (_| |
#   |_| \_|____/_/   \_\ |____/ \___|_|  \__,_| .__/|_|_| |_|\__, |
#                                             |_|            |___/ 

# On importe les bibliothèques nécessaires pour notre programme.
import requests              
import os
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from colorama import Fore, Style

# On définit l'url de base qui est la source de nos données. On ajoutera plus tard l'année spécifique.
url_de_base = "https://www.basketball-reference.com/leagues/NBA_{}_per_game.html"

# On crée un dossier pour stocker notre output si celui-ci n'existe pas déjà.
output_folder = "Statistiques"
os.makedirs(output_folder, exist_ok=True)
graphiques_folder = "Graphiques"
os.makedirs(graphiques_folder, exist_ok=True)

# On défini deux couleurs, light et dark, qu'on utilisera pour colorier les cellules dans notre fichier excel.
fill_light = PatternFill(start_color='FABF8F',
                end_color='FABF8F',
                fill_type='solid')
fill_dark = PatternFill(start_color='FCD3B2',
                end_color='FCD3B2',
                fill_type='solid')

# Ces dictionnaires vides vont stocker les moyennes pour chaque année. On les remplira plus tard.
average_points_per_year = {}
average_age_per_year = {}
average_blocks_per_year = {}
average_assists_per_year = {}
average_two_points_per_year = {}
average_three_points_per_year = {}
average_steals_per_year = {}
average_personal_fouls_per_year = {}
average_games_per_year = {}
average_rebounds_per_year = {}

# On demande à l'utilisateur de saisir l'année de début et la fin de l'intervalle.
annee_debut = int(input("Entrez l'année de début : "))
annee_fin = int(input("Entrez l'année de fin : "))

print("")
print(f"{Fore.WHITE}[{Fore.GREEN}🚧{Fore.WHITE}] Créations des fichiers {Fore.GREEN}excel {Fore.WHITE}en cours.")
print("")

# On parcourt chaque année de l'intervalle.
for annee in range(annee_debut, annee_fin):
    # On forme l'url final en ajoutant l'année à notre url de base.
    url = url_de_base.format(annee)
    # On envoie une requête GET à l'url et on stocke la réponse.
    reponse = requests.get(url)
    # On créé une soup BeautifulSoup de notre réponse pour pouvoir la parser.
    soupe = BeautifulSoup(reponse.text, 'html.parser')
    # On cherche le tableau contenant les données qui nous intéressent.
    tableau = soupe.find('table', {'id': 'per_game_stats'})

    # Pour obtenir le format 23/24 à partir de l'année 2024
    annee_courte = "{}_{}".format(str(annee - 1), str(annee))

    rows_list = []
    try:
        entete = [th.text for th in tableau.thead.findAll('th')]
    except AttributeError:
        print("Vous etes ban IP du site")
    rangees = tableau.tbody.findAll('tr', class_='full_table')
    for rangee in rangees:
        donnees = [td.text for td in rangee.findAll(('td', 'th'))]
        rows_list.append(donnees)

    df = pd.DataFrame(rows_list, columns=entete)

    # Nous convertissons les colonnes des statistiques en données numériques (floats).
    # Ce sont les colonnes dont nous voulons calculer la moyenne plus tard.
    df['PTS'] = df['PTS'].astype(float)
    df['Age'] = df['Age'].astype(float)
    df['BLK'] = pd.to_numeric(df['BLK'], errors='coerce')  # Convertir en float, en transformant les erreurs en NaN
    df['BLK'] = df['BLK'].fillna(0)  # Remplacer les NaN par 0 (sinon crash)
    df['AST'] = df['AST'].astype(float)
    df['2P'] = df['2P'].astype(float)
    df['3P'] = pd.to_numeric(df['3P'], errors='coerce')
    df['3P'] = df['BLK'].fillna(0)
    df['STL'] = pd.to_numeric(df['STL'], errors='coerce')
    df['STL'] = df['BLK'].fillna(0)
    df['PF'] = df['PF'].astype(float)
    df['G'] = df['G'].astype(float)
    df['TRB'] = pd.to_numeric(df['TRB'], errors='coerce')
    df['TRB'] = df['BLK'].fillna(0)

    # Pour chaque statistique, nous calculons la moyenne pour l'année et la stockons dans le dictionnaire correspondant.
    average_points_per_year[annee] = df['PTS'].mean()
    average_age_per_year[annee] = df['Age'].mean()
    average_blocks_per_year[annee] = df['BLK'].mean()
    average_assists_per_year[annee] = df['AST'].mean()
    average_two_points_per_year[annee] = df['2P'].mean()
    average_three_points_per_year[annee] = df['3P'].mean()
    average_steals_per_year[annee] = df['STL'].mean()
    average_personal_fouls_per_year[annee] = df['PF'].mean()
    average_games_per_year[annee] = df['G'].mean()
    average_rebounds_per_year[annee] = df['TRB'].mean()

    # We construct the excel filename using the year and save it into our output_folder.
    nom_fichier_excel = f"Stats_Joueurs_Nba_Année_{annee_courte}.xlsx"
    nom_fichier_excel = os.path.join(output_folder, nom_fichier_excel)

    # Enregistrement du DataFrame en fichier Excel
    df.to_excel(nom_fichier_excel, index=False)

    # Open the excel file with openpyxl to adjust cell alignment and colors
    wb = load_workbook(filename=nom_fichier_excel)
    ws = wb.active

    # Centrer les données et ajuster la couleur de fond de la colonne 'A'
    for cell in ws["A"]:
        cell.alignment = Alignment(horizontal='center')
        cell.fill = fill_dark
        ws.column_dimensions[get_column_letter(1)].width = 3

    # Centrer les données et ajuster la couleur de fond de la colonne 'B'
    for cell in ws["B"]:
        cell.alignment = Alignment(horizontal='center')
        cell.fill = fill_light
        ws.column_dimensions[get_column_letter(2)].width = 25

    for col in range(3, 31):
        col_letter = get_column_letter(col)
        fill_color = fill_light if (col % 2 == 0) else fill_dark

        for cell in ws[col_letter]:
            cell.alignment = Alignment(horizontal='center')
            cell.fill = fill_color

        ws.column_dimensions[col_letter].width = 5

    wb.save(nom_fichier_excel)

    print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Les statistiques des joueurs de l'année {Fore.YELLOW}{annee_courte} {Style.RESET_ALL}ont été sauvegardées dans {Fore.CYAN}{nom_fichier_excel}{Style.RESET_ALL}")
    
print("")
print(f"{Fore.WHITE}[{Fore.GREEN}🚧{Fore.WHITE}] Sauvegarde du {Fore.GREEN}graphique {Fore.WHITE}en cours.{Style.RESET_ALL}")
print("")

# Nous créons un graphique pour chaque dictionnaire de moyennes que nous avons calculées auparavant.
# Chaque clé du dictionnaire est une année, tandis que chaque valeur est la moyenne des statistiques pour cette année.
# Cela nous permet de tracer l'évolution des différentes statistiques au fil du temps.
# Les graphiques sont enregistrés dans le fichier "Graphiques".
plt.figure(figsize=[14,7])
plt.plot(list(average_points_per_year.keys()), list(average_points_per_year.values()), marker='o')
plt.title('Moyenne des points des joueurs de la NBA par année')
plt.xlabel('Année')
plt.ylabel('Points moyens')
plt.grid(True)
plt.savefig(os.path.join(graphiques_folder, 'Moyenne_points_par_annee.png'))
print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Graphique sauvegardé dans {Fore.CYAN}{os.path.join(graphiques_folder, 'Moyenne_points_par_annee.png')}{Style.RESET_ALL}")
plt.close()

plt.figure(figsize=[14,7])
plt.plot(list(average_age_per_year.keys()), list(average_age_per_year.values()), marker='o')
plt.title('Age moyen des joueurs par année')
plt.xlabel('Année')
plt.ylabel('Age moyen')
plt.grid(True)
plt.savefig(os.path.join(graphiques_folder, 'Moyenne_age_par_annee.png'))
print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Graphique sauvegardé dans {Fore.CYAN}{os.path.join(graphiques_folder, 'Moyenne_age_par_annee.png')}{Style.RESET_ALL}")
plt.close()

plt.figure(figsize=[14,7])
plt.plot(list(average_blocks_per_year.keys()), list(average_blocks_per_year.values()), marker='o')
plt.title('Nombre de blocks moyen par année')
plt.xlabel('Année')
plt.ylabel('Blocks moyens')
plt.grid(True)
plt.savefig(os.path.join(graphiques_folder, 'Moyenne_blocks_par_annee.png'))
print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Graphique sauvegardé dans {Fore.CYAN}{os.path.join(graphiques_folder, 'Moyenne_blocks_par_annee.png')}{Style.RESET_ALL}")
plt.close()

plt.figure(figsize=[14,7])
plt.plot(list(average_assists_per_year.keys()), list(average_assists_per_year.values()), marker='o')
plt.title('Nombre de passes décisives moyen par année')
plt.xlabel('Année')
plt.ylabel('Passes décisives moyens')
plt.grid(True)
plt.savefig(os.path.join(graphiques_folder, 'Moyenne_passeD_par_annee.png'))
print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Graphique sauvegardé dans {Fore.CYAN}{os.path.join(graphiques_folder, 'Moyenne_passeD_par_annee.png')}{Style.RESET_ALL}")
plt.close()

plt.figure(figsize=[14,7])
plt.plot(list(average_two_points_per_year.keys()), list(average_two_points_per_year.values()), marker='o')
plt.title('Nombre de deux points moyen par année')
plt.xlabel('Année')
plt.ylabel('Deux points moyens')
plt.grid(True)
plt.savefig(os.path.join(graphiques_folder, 'Moyenne_2P_par_annee.png'))
print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Graphique sauvegardé dans {Fore.CYAN}{os.path.join(graphiques_folder, 'Moyenne_2P_par_annee.png')}{Style.RESET_ALL}")
plt.close()

plt.figure(figsize=[14,7])
plt.plot(list(average_three_points_per_year.keys()), list(average_three_points_per_year.values()), marker='o')
plt.title('Nombre de trois points moyen par année')
plt.xlabel('Année')
plt.ylabel('Trois points moyens')
plt.grid(True)
plt.savefig(os.path.join(graphiques_folder, 'Moyenne_3P_par_annee.png'))
print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Graphique sauvegardé dans {Fore.CYAN}{os.path.join(graphiques_folder, 'Moyenne_3P_par_annee.png')}{Style.RESET_ALL}")
plt.close()

plt.figure(figsize=[14,7])
plt.plot(list(average_steals_per_year.keys()), list(average_steals_per_year.values()), marker='o')
plt.title('Nombre d\'interception moyen par année')
plt.xlabel('Année')
plt.ylabel('Interception moyens')
plt.grid(True)
plt.savefig(os.path.join(graphiques_folder, 'Moyenne_interception_par_annee.png'))
print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Graphique sauvegardé dans {Fore.CYAN}{os.path.join(graphiques_folder, 'Moyenne_interception_par_annee.png')}{Style.RESET_ALL}")
plt.close()

plt.figure(figsize=[14,7])
plt.plot(list(average_personal_fouls_per_year.keys()), list(average_personal_fouls_per_year.values()), marker='o')
plt.title('Nombre de fautes personnels moyen par année')
plt.xlabel('Année')
plt.ylabel('Faute personnel moyens')
plt.grid(True)
plt.savefig(os.path.join(graphiques_folder, 'Moyenne_fautesP_par_annee.png'))
print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Graphique sauvegardé dans {Fore.CYAN}{os.path.join(graphiques_folder, 'Moyenne_fautesP_par_annee.png')}{Style.RESET_ALL}")
plt.close()

plt.figure(figsize=[14,7])
plt.plot(list(average_rebounds_per_year.keys()), list(average_rebounds_per_year.values()), marker='o')
plt.title('Nombre de rebonds moyen par année')
plt.xlabel('Année')
plt.ylabel('Rebonds moyens')
plt.grid(True)
plt.savefig(os.path.join(graphiques_folder, 'Moyenne_rebonds_par_annee.png'))
print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Graphique sauvegardé dans {Fore.CYAN}{os.path.join(graphiques_folder, 'Moyenne_rebonds_par_annee.png')}{Style.RESET_ALL}")
plt.close()

plt.figure(figsize=[14,7])
plt.plot(list(average_games_per_year.keys()), list(average_games_per_year.values()), marker='o')
plt.title('Nombre de match moyen par année')
plt.xlabel('Année')
plt.ylabel('Match moyens')
plt.grid(True)
plt.savefig(os.path.join(graphiques_folder, 'Moyenne_matchs_par_annee.png'))
print(f"{Fore.WHITE}[{Fore.GREEN}+{Fore.WHITE}] Graphique sauvegardé dans {Fore.CYAN}{os.path.join(graphiques_folder, 'Moyenne_matchs_par_annee.png')}{Style.RESET_ALL}")
plt.close()